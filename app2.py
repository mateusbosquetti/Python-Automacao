import openpyxl
import pyautogui

# Carregar o arquivo Excel
workbook = openpyxl.load_workbook('EspacoDoCoordenador.xlsx')
setores_sheet = workbook['setores']

 # Iterar pelas linhas da planilha de setores, começando da segunda linha
for linha in setores_sheet.iter_rows(min_row=2):
    
     # Verificar se linha[0] tem valor
     if linha[0].value:
         # Clicar em "add setor"
         pyautogui.click(921, 287, duration=0.5)

         # Se linha[0] tem valor, selecionar e escrever texto do setor
         pyautogui.click(943, 305, duration=0.5)
         pyautogui.write(str(linha[0].value))

         # Clicar em "Enviar"
         pyautogui.click(936, 351, duration=0.5)

professores_sheet = workbook['professores']

for linha in professores_sheet.iter_rows(min_row=2):
    # Verificar se o valor da coluna 2 é 'TI'
    if linha[2].value == 'TI':
        # Entrar no setor
        pyautogui.click(1241, 492, duration=0.5)

        # Adicionar professor
        pyautogui.click(930, 295, duration=0.5)

        # Nome
        pyautogui.click(919, 305, duration=0.5)
        pyautogui.write(str(linha[0].value))

        # Email
        pyautogui.click(914, 374, duration=0.5)
        pyautogui.write(str(linha[1].value))

        # Confirmar Adição do professor
        pyautogui.click(915, 443, duration=0.5)

        # Home
        pyautogui.click(1706,128, duration=0.5) 

         # Ver Setores
        pyautogui.click(906,319, duration=0.5)

    if linha[2].value == 'Ensino Medio':
        # Entrar no setor
        pyautogui.click(1236,422, duration=0.5)

        # Adicionar professor
        pyautogui.click(930, 295, duration=0.5)

        # Nome
        pyautogui.click(919, 305, duration=0.5)
        pyautogui.write(str(linha[0].value))

        # Email
        pyautogui.click(914, 374, duration=0.5)
        pyautogui.write(str(linha[1].value))

        # Confirmar Adição do professor
        pyautogui.click(915, 443, duration=0.5)

        # Home
        pyautogui.click(1706,128, duration=0.5) 

         # Ver Setores
        pyautogui.click(906,319, duration=0.5)    